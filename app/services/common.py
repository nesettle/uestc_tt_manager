from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

from app.domain import DuplicateRow, FormRecord, QualificationRecord
from app.settings import QUALIFICATION_MASTER_PATH, RUNS_DIR


def collapse_spaces(value: str) -> str:
    return " ".join(str(value or "").split())


def normalize_header(value: Any) -> str:
    return collapse_spaces(str(value or "").strip()).replace("（", "(").replace("）", ")")


def normalize_name(value: Any) -> str:
    text = collapse_spaces(str(value or "").strip())
    while True:
        updated = re.sub(r"\s*[（(][^()（）]*[）)]\s*$", "", text)
        if updated == text:
            break
        text = updated.strip()
    return text


def normalize_college(value: Any) -> str:
    text = collapse_spaces(str(value or "").strip())
    return text.replace("（", "(").replace("）", ")")


def normalize_qq(value: Any) -> str:
    text = str(value or "").strip()
    return "".join(ch for ch in text if ch.isdigit())


def parse_timestamp(value: str) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    for candidate in (text, text.replace("/", "-")):
        try:
            dt = datetime.fromisoformat(candidate)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            continue
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def parse_user_datetime(value: str) -> datetime:
    text = value.strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return datetime.fromisoformat(text).replace(tzinfo=timezone.utc)
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    dt = datetime.fromisoformat(text)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def primary_key(normalized_qq: str, normalized_name: str, normalized_college: str) -> tuple[str, str]:
    if normalized_qq:
        return ("qq", normalized_qq)
    return ("name_college", f"{normalized_name}|{normalized_college}")


def ensure_run_dir(prefix: str) -> Path:
    RUNS_DIR.mkdir(parents=True, exist_ok=True)
    run_dir = RUNS_DIR / f"{prefix}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir


def write_csv(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    rows = list(rows)
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_excel_rows(path: Path) -> list[list[str]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value).strip() for value in row])
    return rows


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [row for row in csv.reader(fh)]


def detect_qualification_columns(headers: list[str]) -> tuple[int, int, int]:
    name_idx = -1
    college_idx = -1
    qq_idx = -1
    normalized_headers = [normalize_header(item) for item in headers]
    for idx, header in enumerate(normalized_headers):
        if name_idx < 0 and header == "姓名":
            name_idx = idx
        if college_idx < 0 and "学院" in header:
            college_idx = idx
        if qq_idx < 0 and header == "QQ号":
            qq_idx = idx
    if name_idx < 0 or college_idx < 0 or qq_idx < 0:
        raise RuntimeError("资格名单表头解析失败，必须包含 姓名 / QQ号 / 包含“学院”的列")
    return name_idx, college_idx, qq_idx


def load_qualification_records(path: Path) -> list[QualificationRecord]:
    workbook = load_workbook(path, data_only=True)
    records: list[QualificationRecord] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [collapse_spaces(str(cell or "").strip()) for cell in rows[0]]
        if not any(headers):
            continue
        name_idx, college_idx, qq_idx = detect_qualification_columns(headers)
        for row_number, row in enumerate(rows[1:], start=2):
            values = list(row)
            name = str(values[name_idx] or "").strip() if name_idx < len(values) else ""
            college = str(values[college_idx] or "").strip() if college_idx < len(values) else ""
            qq = str(values[qq_idx] or "").strip() if qq_idx < len(values) else ""
            if not any((name, college, qq)):
                continue
            records.append(
                QualificationRecord(
                    source=str(path),
                    sheet=sheet.title,
                    row_number=row_number,
                    name=name,
                    college=college,
                    qq=qq,
                    normalized_name=normalize_name(name),
                    normalized_college=normalize_college(college),
                    normalized_qq=normalize_qq(qq),
                )
            )
    return records


EXPORT_ALIASES = {
    "name": ["姓名", "你的姓名", "队员姓名"],
    "college": ["学院", "学院/部门", "学院部门", "学院部门（教学科研单位、研究机构）"],
    "qq": ["QQ号", "QQ", "qq号", "QQ号码"],
    "serial": ["序号"],
    "created_at": ["提交时间", "创建时间"],
    "updated_at": ["更新时间", "更新于"],
}


def detect_export_columns(headers: list[str]) -> dict[str, int]:
    normalized_headers = [normalize_header(item) for item in headers]
    normalized_aliases = {
        key: {normalize_header(alias) for alias in aliases}
        for key, aliases in EXPORT_ALIASES.items()
    }
    found = {key: -1 for key in EXPORT_ALIASES}
    for idx, header in enumerate(normalized_headers):
        for key, aliases in normalized_aliases.items():
            if found[key] >= 0:
                continue
            if header in aliases:
                found[key] = idx
                break
    if found["name"] < 0 or found["college"] < 0 or found["qq"] < 0:
        raise RuntimeError("报名表导出缺少 姓名 / 学院 / QQ号 列，请检查金数据表头")
    return found


def load_form_records_from_export(path: Path, created_after: datetime | None = None) -> list[FormRecord]:
    rows = read_excel_rows(path) if path.suffix.lower() == ".xlsx" else read_csv_rows(path)
    if not rows:
        raise RuntimeError(f"导出文件为空: {path}")
    column_map = detect_export_columns(rows[0])
    results: list[FormRecord] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        name = row[column_map["name"]] if column_map["name"] < len(row) else ""
        college = row[column_map["college"]] if column_map["college"] < len(row) else ""
        qq = row[column_map["qq"]] if column_map["qq"] < len(row) else ""
        created_at = row[column_map["created_at"]] if 0 <= column_map["created_at"] < len(row) else ""
        updated_at = row[column_map["updated_at"]] if 0 <= column_map["updated_at"] < len(row) else ""
        serial_number = row[column_map["serial"]] if 0 <= column_map["serial"] < len(row) else ""
        created_dt = parse_timestamp(created_at)
        if created_after and created_dt and created_dt < created_after:
            continue
        record = FormRecord(
            source=str(path),
            row_number=row_number,
            created_at=created_at,
            updated_at=updated_at,
            serial_number=serial_number,
            name=name,
            college=college,
            qq=qq,
            normalized_name=normalize_name(name),
            normalized_college=normalize_college(college),
            normalized_qq=normalize_qq(qq),
        )
        if record.normalized_name or record.normalized_college or record.normalized_qq:
            results.append(record)
    return results


def dedupe_qualifications(records: list[QualificationRecord]) -> tuple[list[QualificationRecord], list[DuplicateRow]]:
    groups: dict[tuple[str, str], list[QualificationRecord]] = defaultdict(list)
    for record in records:
        groups[primary_key(record.normalized_qq, record.normalized_name, record.normalized_college)].append(record)
    deduped: list[QualificationRecord] = []
    duplicates: list[DuplicateRow] = []
    for (key_type, match_key), items in groups.items():
        deduped.append(items[0])
        if len(items) > 1:
            for item in items:
                duplicates.append(
                    DuplicateRow(
                        source=f"{item.source}#{item.sheet}:{item.row_number}",
                        key_type=key_type,
                        match_key=match_key,
                        name=item.name,
                        college=item.college,
                        qq=item.qq,
                        details="资格名单存在重复匹配键",
                    )
                )
    return deduped, duplicates


def dedupe_form_records(records: list[FormRecord]) -> tuple[list[FormRecord], list[DuplicateRow]]:
    def sort_key(item: FormRecord) -> tuple[datetime, int, int]:
        updated = parse_timestamp(item.updated_at) or parse_timestamp(item.created_at)
        if updated is None:
            updated = datetime.min.replace(tzinfo=timezone.utc)
        serial = int(item.serial_number) if item.serial_number.isdigit() else -1
        return (updated, serial, -item.row_number)

    groups: dict[tuple[str, str], list[FormRecord]] = defaultdict(list)
    for record in records:
        groups[primary_key(record.normalized_qq, record.normalized_name, record.normalized_college)].append(record)
    deduped: list[FormRecord] = []
    duplicates: list[DuplicateRow] = []
    for (key_type, match_key), items in groups.items():
        ordered = sorted(items, key=sort_key, reverse=True)
        deduped.append(ordered[0])
        if len(ordered) > 1:
            for item in ordered:
                duplicates.append(
                    DuplicateRow(
                        source=f"{item.source}#{item.row_number}",
                        key_type=key_type,
                        match_key=match_key,
                        name=item.name,
                        college=item.college,
                        qq=item.qq,
                        details="报名表存在重复提交，主比较仅保留最新一条",
                    )
                )
    return deduped, duplicates


def compare_records(
    qualifications: list[QualificationRecord],
    form_records: list[FormRecord],
) -> tuple[list[dict[str, str]], list[QualificationRecord], list[FormRecord]]:
    remaining_qual = qualifications[:]
    remaining_form = form_records[:]
    matched: list[dict[str, str]] = []

    qual_by_qq = {item.normalized_qq: item for item in remaining_qual if item.normalized_qq}
    form_by_qq = {item.normalized_qq: item for item in remaining_form if item.normalized_qq}
    matched_qual_ids: set[int] = set()
    matched_form_ids: set[int] = set()

    for qq, qual in qual_by_qq.items():
        form = form_by_qq.get(qq)
        if not form:
            continue
        matched_qual_ids.add(id(qual))
        matched_form_ids.add(id(form))
        matched.append(
            {
                "match_type": "qq",
                "qualification_name": qual.name,
                "qualification_college": qual.college,
                "qualification_qq": qual.qq,
                "qualification_source": f"{qual.sheet}:{qual.row_number}",
                "form_name": form.name,
                "form_college": form.college,
                "form_qq": form.qq,
                "form_source": f"{Path(form.source).name}:{form.row_number}",
                "form_created_at": form.created_at,
                "form_updated_at": form.updated_at,
            }
        )

    remaining_qual = [item for item in remaining_qual if id(item) not in matched_qual_ids]
    remaining_form = [item for item in remaining_form if id(item) not in matched_form_ids]

    form_name_map = {
        (item.normalized_name, item.normalized_college): item
        for item in remaining_form
        if item.normalized_name and item.normalized_college
    }
    fallback_form_ids: set[int] = set()
    fallback_qual_ids: set[int] = set()
    for qual in remaining_qual:
        key = (qual.normalized_name, qual.normalized_college)
        if not all(key):
            continue
        form = form_name_map.get(key)
        if not form or id(form) in fallback_form_ids:
            continue
        if qual.normalized_qq and form.normalized_qq:
            continue
        fallback_form_ids.add(id(form))
        fallback_qual_ids.add(id(qual))
        matched.append(
            {
                "match_type": "name_college_fallback",
                "qualification_name": qual.name,
                "qualification_college": qual.college,
                "qualification_qq": qual.qq,
                "qualification_source": f"{qual.sheet}:{qual.row_number}",
                "form_name": form.name,
                "form_college": form.college,
                "form_qq": form.qq,
                "form_source": f"{Path(form.source).name}:{form.row_number}",
                "form_created_at": form.created_at,
                "form_updated_at": form.updated_at,
            }
        )

    remaining_qual = [item for item in remaining_qual if id(item) not in fallback_qual_ids]
    remaining_form = [item for item in remaining_form if id(item) not in fallback_form_ids]
    return matched, remaining_qual, remaining_form


def qualification_record_to_row(record: QualificationRecord) -> dict[str, str]:
    return {
        "name": record.name,
        "college": record.college,
        "qq": record.qq,
        "source": f"{record.sheet}:{record.row_number}",
    }


def form_record_to_row(record: FormRecord) -> dict[str, str]:
    return {
        "name": record.name,
        "college": record.college,
        "qq": record.qq,
        "source": f"{Path(record.source).name}:{record.row_number}",
        "created_at": record.created_at,
        "updated_at": record.updated_at,
    }


def duplicate_row_to_row(record: DuplicateRow) -> dict[str, str]:
    return record.to_dict()


def build_match_key(*, normalized_qq: str, normalized_name: str, normalized_college: str) -> str:
    key_type, key_value = primary_key(normalized_qq, normalized_name, normalized_college)
    return f"{key_type}:{key_value}"


def qualification_record_to_compare_row(record: QualificationRecord, reason: str) -> dict[str, str]:
    return {
        "name": record.name,
        "college": record.college,
        "qq": record.qq,
        "source": f"{record.sheet}:{record.row_number}",
        "match_key": build_match_key(
            normalized_qq=record.normalized_qq,
            normalized_name=record.normalized_name,
            normalized_college=record.normalized_college,
        ),
        "reason": reason,
    }


def form_record_to_compare_row(record: FormRecord, reason: str) -> dict[str, str]:
    return {
        "name": record.name,
        "college": record.college,
        "qq": record.qq,
        "source": f"{Path(record.source).name}:{record.row_number}",
        "match_key": build_match_key(
            normalized_qq=record.normalized_qq,
            normalized_name=record.normalized_name,
            normalized_college=record.normalized_college,
        ),
        "reason": reason,
        "created_at": record.created_at,
        "updated_at": record.updated_at,
    }


def ensure_qualification_master() -> Path:
    if QUALIFICATION_MASTER_PATH.exists():
        return QUALIFICATION_MASTER_PATH
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "资格名单"
    sheet.append(["姓名", "学院", "QQ号", "来源表单", "资格确认时间", "最近更新时间"])
    workbook.save(QUALIFICATION_MASTER_PATH)
    return QUALIFICATION_MASTER_PATH
