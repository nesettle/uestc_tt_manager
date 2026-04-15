from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import shutil

from openpyxl import Workbook, load_workbook

from app.domain import FormRecord, QualificationRecord
from app.services import jinshuju
from app.services.common import (
    ensure_qualification_master,
    ensure_run_dir,
    load_form_records_from_export,
    load_qualification_records,
    normalize_college,
    normalize_name,
    normalize_qq,
    write_json,
)
from app.settings import AppConfig, QUALIFICATION_MASTER_PATH
from app.tasks import TaskContext


MASTER_HEADERS = ["姓名", "学院", "QQ号", "来源表单", "资格确认时间", "最近更新时间"]


@dataclass
class MasterRow:
    name: str
    college: str
    qq: str
    source_form: str
    qualified_at: str
    updated_at: str
    row_number: int | None = None

    @property
    def normalized_name(self) -> str:
        return normalize_name(self.name)

    @property
    def normalized_college(self) -> str:
        return normalize_college(self.college)

    @property
    def normalized_qq(self) -> str:
        return normalize_qq(self.qq)

    def to_dict(self) -> dict[str, str | int]:
        return {
            "row_number": self.row_number or "",
            "姓名": self.name,
            "学院": self.college,
            "QQ号": self.qq,
            "来源表单": self.source_form,
            "资格确认时间": self.qualified_at,
            "最近更新时间": self.updated_at,
        }


def load_master_rows(path: Path = QUALIFICATION_MASTER_PATH) -> list[MasterRow]:
    ensure_qualification_master()
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    results: list[MasterRow] = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row or [])
        if not any(values):
            continue
        while len(values) < len(MASTER_HEADERS):
            values.append("")
        results.append(
            MasterRow(
                name=str(values[0] or "").strip(),
                college=str(values[1] or "").strip(),
                qq=str(values[2] or "").strip(),
                source_form=str(values[3] or "").strip(),
                qualified_at=str(values[4] or "").strip(),
                updated_at=str(values[5] or "").strip(),
                row_number=row_number,
            )
        )
    return results


def save_master_rows(rows: list[MasterRow], path: Path = QUALIFICATION_MASTER_PATH) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "资格名单"
    sheet.append(MASTER_HEADERS)
    for row in rows:
        sheet.append([row.name, row.college, row.qq, row.source_form, row.qualified_at, row.updated_at])
    workbook.save(path)


def serialize_form_record(record: FormRecord) -> dict[str, str]:
    return {
        "name": record.name,
        "college": record.college,
        "qq": record.qq,
        "source": record.source,
        "row_number": str(record.row_number),
        "created_at": record.created_at,
        "updated_at": record.updated_at,
    }


def resolve_names(ctx: TaskContext, config: AppConfig, raw_names: str) -> dict:
    binding = config.qualification_form
    if not binding.title and not binding.entries_url:
        raise RuntimeError("请先在系统配置中绑定资格赛报名表")

    names = [normalize_name(line) for line in raw_names.splitlines()]
    names = [item for item in names if item]
    if not names:
        raise RuntimeError("请至少输入一个姓名")

    run_dir = ensure_run_dir("qualification-resolve")
    export_result = jinshuju.export_form_snapshot(
        ctx,
        config,
        run_dir=run_dir,
        form_title=binding.title,
        entries_url=binding.entries_url,
        export_tag="qualification_form",
    )
    export_path = Path(export_result["export_file"])
    records = load_form_records_from_export(export_path)
    index: dict[str, list[FormRecord]] = {}
    for record in records:
        index.setdefault(record.normalized_name, []).append(record)

    resolved = []
    ambiguous = []
    unmatched = []
    total = len(names)
    for idx, name in enumerate(names, start=1):
        ctx.set_progress(idx, total, f"匹配资格姓名：{name}")
        candidates = index.get(name, [])
        if not candidates:
            unmatched.append(name)
            continue
        candidate_payloads = [serialize_form_record(item) for item in candidates]
        if len(candidate_payloads) == 1:
            resolved.append(candidate_payloads[0])
        else:
            ambiguous.append({"input_name": name, "candidates": candidate_payloads})

    result = {
        "run_dir": str(run_dir),
        "export_file": str(export_path),
        "resolved": resolved,
        "ambiguous": ambiguous,
        "unmatched": unmatched,
    }
    write_json(run_dir / "resolve_result.json", result)
    return result


def apply_updates(ctx: TaskContext, config: AppConfig, selections: list[dict[str, str]]) -> dict:
    if not selections:
        raise RuntimeError("没有可写入的资格名单记录")

    run_dir = ensure_run_dir("qualification-apply")
    ensure_qualification_master()
    if QUALIFICATION_MASTER_PATH.exists():
        shutil.copy2(QUALIFICATION_MASTER_PATH, run_dir / "qualification_snapshot.xlsx")

    now = datetime.now().astimezone().isoformat(timespec="seconds")
    existing = load_master_rows()
    updated = 0
    appended = 0

    for idx, item in enumerate(selections, start=1):
        ctx.set_progress(idx, len(selections), f"写入资格名单：{item.get('name', '')}")
        source_form = config.qualification_form.title or str(item.get("source", "")).strip()
        candidate = MasterRow(
            name=str(item.get("name", "")).strip(),
            college=str(item.get("college", "")).strip(),
            qq=str(item.get("qq", "")).strip(),
            source_form=source_form,
            qualified_at=now,
            updated_at=now,
        )
        match_index = None
        if candidate.normalized_qq:
            for i, current in enumerate(existing):
                if current.normalized_qq and current.normalized_qq == candidate.normalized_qq:
                    match_index = i
                    break
        if match_index is None and candidate.normalized_name and candidate.normalized_college:
            for i, current in enumerate(existing):
                if (
                    current.normalized_name
                    and current.normalized_name == candidate.normalized_name
                    and current.normalized_college == candidate.normalized_college
                    and (not current.normalized_qq or not candidate.normalized_qq)
                ):
                    match_index = i
                    break

        if match_index is None:
            existing.append(candidate)
            appended += 1
            continue

        previous = existing[match_index]
        existing[match_index] = MasterRow(
            name=candidate.name or previous.name,
            college=candidate.college or previous.college,
            qq=candidate.qq or previous.qq,
            source_form=candidate.source_form or previous.source_form,
            qualified_at=previous.qualified_at or candidate.qualified_at,
            updated_at=now,
            row_number=previous.row_number,
        )
        updated += 1

    save_master_rows(existing)
    result = {
        "run_dir": str(run_dir),
        "snapshot_file": str(run_dir / "qualification_snapshot.xlsx"),
        "master_file": str(QUALIFICATION_MASTER_PATH),
        "total_after": len(existing),
        "updated": updated,
        "appended": appended,
    }
    write_json(run_dir / "apply_result.json", result)
    return result


def _extract_selected_row_numbers(selections: list[dict[str, str] | int | str]) -> list[int]:
    row_numbers: list[int] = []
    seen: set[int] = set()
    for item in selections:
        value = item.get("row_number") if isinstance(item, dict) else item
        text = str(value or "").strip()
        if not text.isdigit():
            continue
        row_number = int(text)
        if row_number in seen:
            continue
        seen.add(row_number)
        row_numbers.append(row_number)
    return row_numbers


def delete_rows(ctx: TaskContext, selections: list[dict[str, str] | int | str]) -> dict:
    row_numbers = _extract_selected_row_numbers(selections)
    if not row_numbers:
        raise RuntimeError("请先勾选要删除的资格名单成员")

    run_dir = ensure_run_dir("qualification-delete")
    ensure_qualification_master()
    snapshot_file = run_dir / "qualification_snapshot.xlsx"
    if QUALIFICATION_MASTER_PATH.exists():
        shutil.copy2(QUALIFICATION_MASTER_PATH, snapshot_file)

    existing = load_master_rows()
    selected_set = set(row_numbers)
    deleted_rows: list[dict[str, str | int]] = []
    remaining: list[MasterRow] = []
    total = len(existing)

    for idx, row in enumerate(existing, start=1):
        ctx.set_progress(idx, total, f"检查资格名单：{row.name}")
        if row.row_number in selected_set:
            deleted_rows.append(row.to_dict())
            continue
        remaining.append(row)

    if not deleted_rows:
        raise RuntimeError("未匹配到要删除的资格名单成员")

    save_master_rows(remaining)
    result = {
        "run_dir": str(run_dir),
        "snapshot_file": str(snapshot_file),
        "master_file": str(QUALIFICATION_MASTER_PATH),
        "deleted": len(deleted_rows),
        "total_after": len(remaining),
        "deleted_rows": deleted_rows,
    }
    write_json(run_dir / "delete_result.json", result)
    return result


def current_master_preview() -> dict:
    records = load_master_rows()
    return {
        "master_file": str(QUALIFICATION_MASTER_PATH),
        "count": len(records),
        "rows": [row.to_dict() for row in records],
    }


def export_master_file() -> Path:
    ensure_qualification_master()
    return QUALIFICATION_MASTER_PATH


def load_master_as_qualification_records() -> list[QualificationRecord]:
    ensure_qualification_master()
    return load_qualification_records(QUALIFICATION_MASTER_PATH)
